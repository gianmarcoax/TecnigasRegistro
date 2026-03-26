"""
Servidor local – Catálogo / Recepción Tecnigass
Uso: python server.py
Abre: http://localhost:8080
"""

import xmlrpc.client
import json
import urllib.parse
import os
import sys
import base64
import io
from http.server import HTTPServer, SimpleHTTPRequestHandler

# ── Odoo credentials (from env vars or config.json) ───────────────
_config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
_config = {}

if os.path.exists(_config_file):
    with open(_config_file, 'r', encoding='utf-8') as f:
        _config = json.load(f)
else:
    print("Aviso: 'config.json' no encontrado. Intentando usar variables de entorno.")

ODOO_URL    = os.environ.get('ODOO_URL') or _config.get('ODOO_URL', 'https://tecnigass.pe')
ODOO_DB     = os.environ.get('ODOO_DB') or _config.get('ODOO_DB', 'db_tecnigas')
ODOO_USER   = os.environ.get('ODOO_USER') or _config.get('ODOO_USER', '')
ODOO_APIKEY = os.environ.get('ODOO_APIKEY') or _config.get('ODOO_APIKEY', '')

if not ODOO_USER or not ODOO_APIKEY:
    print("Error: Faltan credenciales ODOO_USER o ODOO_APIKEY. Revisa config.json o tus variables de entorno.")
    sys.exit(1)

# Plantilla Excel (opcional, solo se usa si existe localmente)
TEMPLATE_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), '010.xlsx')

_uid = None

def get_uid():
    global _uid
    if _uid:
        return _uid
    common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
    _uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_APIKEY, {})
    return _uid

def odoo_call(model, method, args, kwargs):
    uid = get_uid()
    models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
    return models.execute_kw(ODOO_DB, uid, ODOO_APIKEY, model, method, args, kwargs)


# ── Excel builder (en memoria) ────────────────────────────────
def build_excel_bytes(rows):
    """
    Genera el Excel en memoria y devuelve los bytes.
    No guarda nada en disco (compatible con Railway/cloud).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Crear workbook basado en la plantilla si existe localmente
    if os.path.exists(TEMPLATE_XLSX):
        wb = openpyxl.load_workbook(TEMPLATE_XLSX)
        ws = wb.active
        # Limpiar filas de datos (mantener cabecera en fila 1)
        for row_idx in range(ws.max_row, 1, -1):
            ws.delete_rows(row_idx)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Cabecera
        headers = ['Cantidad a la mano', 'Nombre', 'Precio de venta', 'Referencia interna']
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    # Escribir filas de datos
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, row in enumerate(rows, 2):
        # Columna 4: Referencia interna / barcode — PRIMERO, para que el border
        # no sobreescriba el quotePrefix al llamar ws.cell() de nuevo.
        code_val = str(row.get('default_code', '') or '')
        c4 = ws.cell(row=row_idx, column=4, value=code_val)
        c4.number_format = '@'   # formato texto — BarTender lo lee correctamente
        c4.border = border

        c1 = ws.cell(row=row_idx, column=1, value=row.get('tickets', 1))
        c1.border = border
        c2 = ws.cell(row=row_idx, column=2, value=row.get('name', ''))
        c2.border = border
        c3 = ws.cell(row=row_idx, column=3, value=row.get('list_price', 0))
        c3.border = border

    # Ajustar ancho de columnas
    col_widths = [20, 45, 18, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTTP Handler ──────────────────────────────────────────────
class Handler(SimpleHTTPRequestHandler):

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)

        if parsed.path == '/':
            self.send_response(301)
            self.send_header('Location', '/dashboard.html')
            self.end_headers()
            return

        if parsed.path == '/api':
            params = dict(urllib.parse.parse_qsl(parsed.query))
            action = params.get('action', 'search')
            try:
                if action == 'search':
                    result = self.handle_search(params)
                elif action == 'categories':
                    result = self.handle_categories()
                elif action == 'products_by_ids':
                    result = self.handle_products_by_ids(params)
                elif action == 'kit_components':
                    result = self.handle_kit_components(params)
                elif action == 'locations':
                    result = self.handle_locations()
                elif action == 'stock_by_location':
                    result = self.handle_stock_by_location(params)
                elif action == 'history_receptions':
                    result = self.handle_history_receptions(params)
                elif action == 'history_transfers':
                    result = self.handle_history_transfers(params)
                elif action == 'picking_detail':
                    result = self.handle_picking_detail(params)
                elif action == 'next_product_code':
                    result = self.handle_next_product_code(params)
                else:
                    result = {'error': 'Acción no reconocida'}
            except Exception as e:
                result = {'error': str(e)}

            self._json_response(result)
            return

        # Servir archivos estáticos
        super().do_GET()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)

        if parsed.path == '/api':
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length)
            try:
                data = json.loads(body.decode('utf-8'))
                action = data.get('action', '')
                if action == 'export_excel':
                    result = self.handle_export_excel(data)
                elif action == 'receive':
                    result = self.handle_receive(data)
                elif action == 'transfer':
                    result = self.handle_transfer(data)
                elif action == 'revert_picking':
                    result = self.handle_revert_picking(data)
                elif action == 'create_product':
                    result = self.handle_create_product(data)
                elif action == 'update_product':
                    result = self.handle_update_product(data)
                else:
                    result = {'error': 'Acción POST no reconocida'}
            except Exception as e:
                result = {'error': str(e)}

            self._json_response(result)
            return

        self.send_response(405)
        self.end_headers()

    def _json_response(self, result):
        # Añadir cabeceras CORS por si acaso
        body = json.dumps(result, ensure_ascii=False).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    # ── Search products ──────────────────────────────────────
    def handle_search(self, params):
        q     = params.get('q', '').strip()
        categ = int(params.get('categ', 0) or 0)

        domain = [['active', '=', True], ['sale_ok', '=', True]]
        if q:
            domain += ['|', ['name', 'ilike', q], ['default_code', 'ilike', q]]
        if categ:
            domain.append(['categ_id', 'child_of', categ])

        products = odoo_call('product.template', 'search_read', [domain], {
            'fields': ['id', 'name', 'default_code', 'barcode', 'categ_id',
                       'list_price', 'standard_price', 'qty_available',
                       'uom_id', 'image_128'],
            'limit': 120,
            'order': 'default_code asc',
        })

        # Marcar cuáles son kits (tienen mrp.bom tipo phantom)
        if products:
            tmpl_ids = [p['id'] for p in products]
            try:
                boms = odoo_call('mrp.bom', 'search_read',
                    [[['product_tmpl_id', 'in', tmpl_ids], ['type', '=', 'phantom']]],
                    {'fields': ['product_tmpl_id']})
                kit_ids = {b['product_tmpl_id'][0] for b in boms}
            except Exception:
                kit_ids = set()

            for p in products:
                p['is_kit'] = p['id'] in kit_ids

        return {'products': products or []}

    # ── Categories ────────────────────────────────────────────
    def handle_categories(self):
        cats = odoo_call('product.category', 'search_read', [[]], {
            'fields': ['id', 'name', 'complete_name'],
            'order': 'complete_name asc',
        })
        return {'categories': cats or []}

    # ── Full product data by IDs ───────────────────────────────
    def handle_products_by_ids(self, params):
        raw_ids = params.get('ids', '')
        try:
            ids = [int(i) for i in raw_ids.split(',') if i.strip()]
        except ValueError:
            return {'error': 'IDs inválidos'}

        if not ids:
            return {'products': []}

        products = odoo_call('product.template', 'search_read',
            [[['id', 'in', ids]]],
            {
                'fields': ['id', 'name', 'default_code', 'barcode',
                           'list_price', 'standard_price', 'categ_id',
                           'uom_id', 'image_1920'],
                'order': 'categ_id asc, default_code asc',
            })
        return {'products': products or []}

    # ── Kit components ────────────────────────────────────────
    def handle_kit_components(self, params):
        """
        Devuelve los componentes de un kit (mrp.bom tipo phantom)
        para un product.template dado.
        """
        tmpl_id = int(params.get('id', 0) or 0)
        if not tmpl_id:
            return {'error': 'Falta el parámetro id'}

        # Buscar la BoM de tipo phantom para este template
        boms = odoo_call('mrp.bom', 'search_read',
            [[['product_tmpl_id', '=', tmpl_id], ['type', '=', 'phantom']]],
            {'fields': ['id', 'product_tmpl_id', 'bom_line_ids'], 'limit': 1})

        if not boms:
            return {'components': [], 'is_kit': False}

        bom = boms[0]
        line_ids = bom.get('bom_line_ids', [])

        if not line_ids:
            return {'components': [], 'is_kit': True}

        # Obtener detalle de cada línea
        lines = odoo_call('mrp.bom.line', 'search_read',
            [[['id', 'in', line_ids]]],
            {'fields': ['product_id', 'product_qty', 'product_uom_id']})

        # Enriquecer con imagen y datos del producto
        product_ids = [l['product_id'][0] for l in lines if l.get('product_id')]
        products_map = {}
        if product_ids:
            prods = odoo_call('product.product', 'search_read',
                [[['id', 'in', product_ids]]],
                {'fields': ['id', 'name', 'default_code', 'list_price',
                            'standard_price', 'image_128']})
            products_map = {p['id']: p for p in prods}

        components = []
        for line in lines:
            if not line.get('product_id'):
                continue
            pid = line['product_id'][0]
            prod = products_map.get(pid, {})
            components.append({
                'product_id': pid,
                'name': prod.get('name', line['product_id'][1]),
                'default_code': prod.get('default_code', ''),
                'list_price': prod.get('list_price', 0),
                'standard_price': prod.get('standard_price', 0),
                'image_128': prod.get('image_128', False),
                'qty_bom': line.get('product_qty', 1),
                'uom': line.get('product_uom_id', [None, ''])[1] if line.get('product_uom_id') else '',
            })

        return {'components': components, 'is_kit': True, 'bom_id': bom['id']}

    # ── Locations ─────────────────────────────────────────────
    def handle_locations(self):
        """
        Devuelve ubicaciones de stock relevantes para la recepción.
        Prioriza PUN/ALMACEN y PUN/TIENDA.
        """
        locs = odoo_call('stock.location', 'search_read',
            [[['usage', '=', 'internal'], ['active', '=', True]]],
            {
                'fields': ['id', 'name', 'complete_name'],
            })

        # Poner primero las ubicaciones de PUNO si existen
        priority = ['ALMACEN', 'TIENDA']
        def sort_key(loc):
            cn = loc.get('complete_name', '').upper()
            for i, kw in enumerate(priority):
                if kw in cn:
                    return (0, i, cn)
            return (1, 0, cn)

        locs_sorted = sorted(locs or [], key=sort_key)
        return {'locations': locs_sorted}

    # ── Export Excel ──────────────────────────────────────────
    def handle_export_excel(self, data):
        """
        Recibe: { action, rows: [{name, default_code, list_price, tickets}, ...] }
        Genera el Excel en memoria y lo devuelve como base64 para que
        el navegador lo descargue directamente (sin guardar en disco).
        """
        rows = data.get('rows', [])
        if not rows:
            return {'error': 'No hay productos para exportar'}

        try:
            xlsx_bytes = build_excel_bytes(rows)
            b64 = base64.b64encode(xlsx_bytes).decode('utf-8')
            return {
                'ok': True,
                'filename': '010.xlsx',
                'data': b64,
                'count': len(rows),
            }
        except ImportError:
            return {'error': 'Falta openpyxl. Instala con: pip install openpyxl'}
        except Exception as e:
            return {'error': str(e)}

    # ── History: recepciones desde Odoo ──────────────────────────
    def handle_history_receptions(self, params):
        """
        Devuelve los stock.picking de tipo incoming (recepciones).
        GET /api?action=history_receptions&limit=50&state=done
        """
        limit  = int(params.get('limit', 60) or 60)
        state  = params.get('state', '')   # '', 'done', 'confirmed', 'draft'

        domain = [['picking_type_code', '=', 'incoming']]
        if state:
            domain.append(['state', '=', state])

        pickings = odoo_call('stock.picking', 'search_read', [domain], {
            'fields': ['id', 'name', 'date_done', 'date', 'state',
                       'origin', 'partner_id', 'move_ids_without_package'],
            'order':  'date desc',
            'limit':  limit,
        })

        # Contar líneas de cada picking
        for p in (pickings or []):
            p['line_count'] = len(p.get('move_ids_without_package', []))

        return {'pickings': pickings or []}

    # ── History: traslados desde Odoo ────────────────────────────
    def handle_history_transfers(self, params):
        """
        Devuelve los stock.picking de tipo internal (traslados internos).
        GET /api?action=history_transfers&limit=50&state=done
        """
        limit  = int(params.get('limit', 60) or 60)
        state  = params.get('state', '')

        domain = [['picking_type_code', '=', 'internal']]
        if state:
            domain.append(['state', '=', state])

        pickings = odoo_call('stock.picking', 'search_read', [domain], {
            'fields': ['id', 'name', 'date_done', 'date', 'state',
                       'origin', 'partner_id', 'move_ids_without_package',
                       'location_id', 'location_dest_id'],
            'order':  'date desc',
            'limit':  limit,
        })

        for p in (pickings or []):
            p['line_count'] = len(p.get('move_ids_without_package', []))

        return {'pickings': pickings or []}

    # ── Revert picking: cancelar o devolver ──────────────────────
    def handle_revert_picking(self, data):
        """
        Cancela o revierte un picking.
        - draft/confirmed/assigned → button_cancel
        - done → crea un retorno via stock.return.picking
        POST /api  { action: revert_picking, id: <int> }
        """
        picking_id = int(data.get('id', 0) or 0)
        if not picking_id:
            return {'error': 'Falta el id del picking'}

        picks = odoo_call('stock.picking', 'read', [[picking_id]], {
            'fields': ['id', 'state', 'name']
        })
        if not picks:
            return {'error': 'Picking no encontrado'}

        state = picks[0].get('state', '')
        name  = picks[0].get('name', '')

        if state == 'done':
            # Crear retorno: wizard stock.return.picking
            try:
                ctx = odoo_call('stock.picking', 'action_open_return',
                                [[picking_id]], {})
                # Odoo 18: action_open_return devuelve una acción,
                # usamos el wizard id del context si existe.
                # Fallback: crear el wizard directamente.
            except Exception:
                ctx = None

            try:
                wiz_id = odoo_call('stock.return.picking', 'create',
                                   [{'picking_id': picking_id}], {})
                result = odoo_call('stock.return.picking', 'create_returns',
                                   [[wiz_id]], {})
                new_picking_id = None
                if isinstance(result, dict) and result.get('res_id'):
                    new_picking_id = result['res_id']
                return {
                    'ok': True,
                    'action': 'returned',
                    'name': name,
                    'new_picking_id': new_picking_id,
                }
            except Exception as e:
                return {'error': f'No se pudo crear el retorno: {e}'}
        else:
            # Cancelar (sólo si no está hecho)
            try:
                odoo_call('stock.picking', 'action_cancel', [[picking_id]], {})
                return {'ok': True, 'action': 'cancelled', 'name': name}
            except Exception as e:
                return {'error': f'No se pudo cancelar: {e}'}

    # ── Next product code: auto-código único global ───────────────
    def handle_next_product_code(self, params):
        """
        Calcula el siguiente código de producto para una categoría.
        Formato: [CAT3][SUBCAT3][CORR3]  ej: 003031001
        Verifica unicidad global en product.template.
        GET /api?action=next_product_code&categ_id=<int>&complete_name=<str>
        """
        categ_id      = int(params.get('categ_id', 0) or 0)
        complete_name = params.get('complete_name', '')

        if not categ_id or not complete_name:
            return {'error': 'Faltan parámetros categ_id y complete_name'}

        # Extraer partes del nombre "003 UTENSILIOS / 031 TAPERS"
        # o "003 UTENSILIOS" si es categoría raíz
        import re
        parts = [p.strip() for p in complete_name.split('/')]
        def extract_num(s):
            m = re.match(r'^(\d+)', s.strip())
            return m.group(1).zfill(3) if m else '000'

        if len(parts) >= 2:
            cat_num    = extract_num(parts[-2])
            subcat_num = extract_num(parts[-1])
        else:
            cat_num    = extract_num(parts[0])
            subcat_num = '000'

        prefix = cat_num + subcat_num  # ej: "003031"

        # Buscar todos los códigos que empiecen con ese prefijo en Odoo
        existing = odoo_call('product.template', 'search_read',
            [[['default_code', 'like', prefix + '%']]],
            {'fields': ['default_code']})

        max_corr = 0
        pat = re.compile(r'^' + re.escape(prefix) + r'(\d{3})$')
        for prod in (existing or []):
            code = prod.get('default_code') or ''
            m = pat.match(code)
            if m:
                n = int(m.group(1))
                if n > max_corr:
                    max_corr = n

        # Buscar el siguiente que no exista globalmente
        candidate = max_corr + 1
        all_codes = {(p.get('default_code') or '').strip()
                     for p in (existing or [])}
        while candidate <= 999:
            proposed = prefix + str(candidate).zfill(3)
            if proposed not in all_codes:
                break
            candidate += 1

        if candidate > 999:
            return {'error': 'No hay más códigos disponibles para esta categoría'}

        return {
            'code': prefix + str(candidate).zfill(3),
            'prefix': prefix,
            'next_corr': candidate,
        }

    # ── Create product ────────────────────────────────────────────
    def handle_create_product(self, data):
        """
        Crea un product.template en Odoo con los valores recibidos.
        POST { action, name, default_code, categ_id, list_price, image_base64? }
        """
        name         = (data.get('name') or '').strip()
        default_code = (data.get('default_code') or '').strip()
        categ_id     = int(data.get('categ_id', 0) or 0)
        list_price   = float(data.get('list_price', 0) or 0)
        image_b64    = data.get('image_base64') or None

        if not name:
            return {'error': 'El nombre del producto es requerido'}
        if not default_code:
            return {'error': 'La referencia/código es requerida'}
        if not categ_id:
            return {'error': 'La categoría es requerida'}

        # Verificar unicidad global del código
        dupes = odoo_call('product.template', 'search_read',
            [[['default_code', '=', default_code]]],
            {'fields': ['id', 'name'], 'limit': 1})
        if dupes:
            return {'error': f"El código '{default_code}' ya existe: {dupes[0].get('name','')}"}

        # Verificar unicidad del barcode
        dupes_bc = odoo_call('product.template', 'search_read',
            [[['barcode', '=', default_code]]],
            {'fields': ['id', 'name'], 'limit': 1})
        if dupes_bc:
            return {'error': f"El código de barras '{default_code}' ya existe: {dupes_bc[0].get('name','')}"}

        vals = {
            'name':          name,
            'default_code':  default_code,
            'barcode':       default_code,
            'categ_id':      categ_id,
            'list_price':    list_price,
            'type':          'product',   # Bienes (storable)
            'tracking':      'none',
            'sale_ok':       True,
            'purchase_ok':   True,
            'available_in_pos': True,
        }
        if image_b64:
            vals['image_1920'] = image_b64

        try:
            new_id = odoo_call('product.template', 'create', [vals], {})
            return {'ok': True, 'id': new_id, 'name': name, 'code': default_code}
        except Exception as e:
            return {'error': str(e)}

    # ── Update product ────────────────────────────────────────────
    def handle_update_product(self, data):
        """
        Actualiza campos de un product.template existente.
        POST { action, id, name?, list_price?, categ_id?, image_base64? }
        Referencia y barcode siempre se sincronizan entre sí.
        """
        tmpl_id = int(data.get('id', 0) or 0)
        if not tmpl_id:
            return {'error': 'Falta el id del producto'}

        vals = {}
        if 'name'        in data: vals['name']        = data['name']
        if 'list_price'  in data: vals['list_price']  = float(data['list_price'])
        if 'categ_id'    in data: vals['categ_id']    = int(data['categ_id'])
        if 'image_base64' in data and data['image_base64']:
            vals['image_1920'] = data['image_base64']

        if 'default_code' in data:
            new_code = (data['default_code'] or '').strip()
            if new_code:
                # Verificar unicidad excluyendo el propio producto
                dupes = odoo_call('product.template', 'search_read',
                    [[['default_code', '=', new_code],
                      ['id', '!=', tmpl_id]]],
                    {'fields': ['id'], 'limit': 1})
                if dupes:
                    return {'error': f"El código '{new_code}' ya está en uso por otro producto"}
                vals['default_code'] = new_code
                vals['barcode']      = new_code  # siempre sincronizados

        if not vals:
            return {'ok': True, 'message': 'Sin cambios'}

        try:
            odoo_call('product.template', 'write', [[tmpl_id], vals], {})
            return {'ok': True, 'id': tmpl_id}
        except Exception as e:
            return {'error': str(e)}

    # ── Picking detail: líneas de movimiento ──────────────────────
    def handle_picking_detail(self, params):
        """
        Devuelve el detalle completo de un stock.picking: cabecera + líneas.
        GET /api?action=picking_detail&id=<picking_id>
        """
        picking_id = int(params.get('id', 0) or 0)
        if not picking_id:
            return {'error': 'Falta el parámetro id'}

        # Cabecera del picking
        picks = odoo_call('stock.picking', 'read', [[picking_id]], {
            'fields': ['id', 'name', 'date_done', 'date', 'state',
                       'origin', 'partner_id', 'location_dest_id',
                       'move_ids_without_package']
        })
        if not picks:
            return {'error': 'Recepción no encontrada'}
        pick = picks[0]

        move_ids = pick.get('move_ids_without_package', [])
        lines = []
        if move_ids:
            moves = odoo_call('stock.move', 'read', [move_ids], {
                'fields': ['id', 'name', 'product_id', 'product_uom_qty',
                           'quantity', 'product_uom', 'state']
            })
            # Enriquecer con default_code y list_price del template
            tmpl_ids = []
            prod_tmpl_map = {}  # product.product id → template id
            for m in moves:
                if m.get('product_id'):
                    pp_id = m['product_id'][0]
                    tmpl_ids.append(pp_id)

            if tmpl_ids:
                variants = odoo_call('product.product', 'search_read',
                    [[['id', 'in', tmpl_ids]]],
                    {'fields': ['id', 'product_tmpl_id', 'default_code', 'list_price']})
                prod_tmpl_map = {v['id']: v for v in variants}

            for m in moves:
                pp_id = m['product_id'][0] if m.get('product_id') else None
                vdata = prod_tmpl_map.get(pp_id, {})
                lines.append({
                    'move_id':      m['id'],
                    'product_id':   pp_id,
                    'name':         m.get('name', ''),
                    'default_code': vdata.get('default_code', ''),
                    'list_price':   vdata.get('list_price', 0),
                    'qty_done':     m.get('quantity', 0),
                    'qty_ordered':  m.get('product_uom_qty', 0),
                    'uom':          m['product_uom'][1] if m.get('product_uom') else '',
                    'tickets':      int(m.get('quantity', 1)) or 1,
                })

        return {'picking': pick, 'lines': lines}

    def handle_stock_by_location(self, params):
        """
        Devuelve stock (qty_available) de una lista de product.template IDs
        en una ubicación dada.
        GET /api?action=stock_by_location&loc_id=8&tmpl_ids=1,2,3
        """
        loc_id = int(params.get('loc_id', 0) or 0)
        raw_ids = params.get('tmpl_ids', '')
        try:
            tmpl_ids = [int(i) for i in raw_ids.split(',') if i.strip()]
        except ValueError:
            return {'error': 'IDs inválidos'}

        if not loc_id or not tmpl_ids:
            return {'stock': {}}

        # Buscar product.product IDs de esos templates
        prods = odoo_call('product.product', 'search_read',
            [[['product_tmpl_id', 'in', tmpl_ids]]],
            {'fields': ['id', 'product_tmpl_id']})
        prod_to_tmpl = {p['id']: p['product_tmpl_id'][0] for p in prods}
        prod_ids = list(prod_to_tmpl.keys())

        if not prod_ids:
            return {'stock': {}}

        quants = odoo_call('stock.quant', 'search_read',
            [[['location_id', '=', loc_id],
              ['product_id', 'in', prod_ids]]],
            {'fields': ['product_id', 'quantity', 'reserved_quantity']})

        stock = {}  # tmpl_id -> qty disponible
        for q in quants:
            prod_id = q['product_id'][0]
            tmpl_id = prod_to_tmpl.get(prod_id)
            if tmpl_id:
                available = q.get('quantity', 0) - q.get('reserved_quantity', 0)
                stock[tmpl_id] = stock.get(tmpl_id, 0) + available

        return {'stock': stock}

    # ── Receive: crear stock.picking en Odoo ────────────────────
    def handle_receive(self, data):
        """
        Crea un stock.picking de tipo incoming en Odoo y lo valida.
        data: {
          action: 'receive',
          location_dest_id: <int>,   # PUN/ALMACEN o PUN/TIENDA
          rows: [{product_id: int, name: str, qty: float}, ...]
        }
        """
        rows = data.get('rows', [])
        loc_dest_id = int(data.get('location_dest_id', 0) or 0)

        if not rows:
            return {'error': 'No hay productos para recepcionar'}
        if not loc_dest_id:
            return {'error': 'Debes seleccionar un destino (almacén o tienda)'}

        # 1. Encontrar el picking type incoming que tenga ese destino (o el primero incoming del almacén)
        pts = odoo_call('stock.picking.type', 'search_read',
            [[['code', '=', 'incoming'],
              ['default_location_dest_id', '=', loc_dest_id]]],
            {'fields': ['id', 'name', 'default_location_src_id'], 'limit': 1})

        if not pts:
            # Fallback: cualquier picking type incoming
            pts = odoo_call('stock.picking.type', 'search_read',
                [[['code', '=', 'incoming']]],
                {'fields': ['id', 'name', 'default_location_src_id'], 'limit': 1})

        if not pts:
            return {'error': 'No se encontró un tipo de operación de recepción en Odoo'}

        pt = pts[0]
        picking_type_id  = pt['id']
        loc_src_id = pt['default_location_src_id'][0] if pt.get('default_location_src_id') else None

        # Ubicación origen por defecto: proveedor genérico
        if not loc_src_id:
            sup = odoo_call('stock.location', 'search',
                [[['usage', '=', 'supplier']]], {'limit': 1})
            loc_src_id = sup[0] if sup else None

        if not loc_src_id:
            return {'error': 'No se pudo determinar la ubicación de origen (proveedor)'}

        # 2. Crear el picking
        picking_vals = {
            'picking_type_id': picking_type_id,
            'location_id':     loc_src_id,
            'location_dest_id': loc_dest_id,
            'origin': 'Recepción Web',
        }
        picking_id = odoo_call('stock.picking', 'create', [picking_vals], {})

        # 3. Resolver IDs: el frontend envía product.template IDs,
        #    pero stock.move necesita product.product IDs (variante).
        tmpl_ids = list({int(r.get('product_id', 0)) for r in rows if r.get('product_id')})
        tmpl_to_variant = {}
        if tmpl_ids:
            variants = odoo_call('product.product', 'search_read',
                [[['product_tmpl_id', 'in', tmpl_ids], ['active', '=', True]]],
                {'fields': ['id', 'product_tmpl_id', 'uom_id'], 'limit': len(tmpl_ids) * 5})
            for v in variants:
                tid = v['product_tmpl_id'][0]
                if tid not in tmpl_to_variant:
                    tmpl_to_variant[tid] = (v['id'], v['uom_id'][0] if v.get('uom_id') else 1)

        # 4. Crear los movimientos de stock
        for row in rows:
            qty = float(row.get('qty', 1))
            if qty <= 0:
                continue
            tmpl_id = int(row.get('product_id', 0))
            if not tmpl_id:
                continue
            variant_info = tmpl_to_variant.get(tmpl_id)
            if not variant_info:
                print(f'  [RECV] Warning: no variante para template {tmpl_id}')
                continue
            pp_id, uom_id = variant_info
            odoo_call('stock.move', 'create', [{
                'name':             row.get('name', 'Producto'),
                'picking_id':       picking_id,
                'product_id':       pp_id,
                'product_uom_qty':  qty,
                'product_uom':      uom_id,
                'location_id':      loc_src_id,
                'location_dest_id': loc_dest_id,
            }], {})

        # 4. Confirmar y validar la recepción
        odoo_call('stock.picking', 'action_confirm', [[picking_id]], {})
        # Asignar cantidades disponibles
        odoo_call('stock.picking', 'action_assign', [[picking_id]], {})
        # Validar (marcar como Hecho)
        try:
            odoo_call('stock.picking', 'button_validate', [[picking_id]], {})
            state = 'done'
        except Exception as e:
            state = 'confirmed'
            print(f'  [RECV] Validación automática no completada: {e}')

        # Obtener el nombre del picking creado
        pick_data = odoo_call('stock.picking', 'read', [[picking_id]], {'fields': ['name', 'state']})
        pick_name = pick_data[0]['name'] if pick_data else str(picking_id)

        return {
            'ok': True,
            'picking_id': picking_id,
            'picking_name': pick_name,
            'state': state,
            'count': len(rows),
        }

    # ── Transfer: movimiento interno en Odoo ────────────────────
    def handle_transfer(self, data):
        """
        Crea un stock.picking de tipo 'internal' en Odoo (traslado entre ubicaciones).
        data: {
          action: 'transfer',
          location_src_id: <int>,    # PUN/ALMACEN o PUN/TIENDA (origen)
          location_dest_id: <int>,   # PUN/TIENDA o PUN/ALMACEN (destino)
          rows: [{product_id: int, name: str, qty: float}, ...]
        }
        """
        rows         = data.get('rows', [])
        loc_src_id   = int(data.get('location_src_id',  0) or 0)
        loc_dest_id  = int(data.get('location_dest_id', 0) or 0)

        if not rows:        return {'error': 'No hay productos para trasladar'}
        if not loc_src_id:  return {'error': 'Selecciona la ubicación de origen'}
        if not loc_dest_id: return {'error': 'Selecciona la ubicación de destino'}
        if loc_src_id == loc_dest_id:
            return {'error': 'Origen y destino deben ser diferentes'}

        # 1. Buscar picking type internal
        pts = odoo_call('stock.picking.type', 'search_read',
            [[['code', '=', 'internal'],
              ['default_location_src_id', '=', loc_src_id]]],
            {'fields': ['id', 'name'], 'limit': 1})
        if not pts:
            pts = odoo_call('stock.picking.type', 'search_read',
                [[['code', '=', 'internal']]],
                {'fields': ['id', 'name'], 'limit': 1})
        if not pts:
            return {'error': 'No se encontró tipo de operación de traslado interno en Odoo'}

        picking_type_id = pts[0]['id']

        # 2. Crear el picking
        picking_id = odoo_call('stock.picking', 'create', [{
            'picking_type_id':  picking_type_id,
            'location_id':      loc_src_id,
            'location_dest_id': loc_dest_id,
            'origin':           'Traslado Web',
        }], {})

        # 3. Resolver template IDs → product.product IDs
        tmpl_ids = list({int(r.get('product_id', 0)) for r in rows if r.get('product_id')})
        tmpl_to_variant = {}
        if tmpl_ids:
            variants = odoo_call('product.product', 'search_read',
                [[['product_tmpl_id', 'in', tmpl_ids], ['active', '=', True]]],
                {'fields': ['id', 'product_tmpl_id', 'uom_id'], 'limit': len(tmpl_ids) * 5})
            for v in variants:
                tid = v['product_tmpl_id'][0]
                if tid not in tmpl_to_variant:
                    tmpl_to_variant[tid] = (v['id'], v['uom_id'][0] if v.get('uom_id') else 1)

        # 4. Crear movimientos de stock
        for row in rows:
            qty = float(row.get('qty', 1))
            if qty <= 0: continue
            tmpl_id = int(row.get('product_id', 0))
            if not tmpl_id: continue
            variant_info = tmpl_to_variant.get(tmpl_id)
            if not variant_info:
                print(f'  [TRNF] Warning: no variante para template {tmpl_id}')
                continue
            pp_id, uom_id = variant_info
            odoo_call('stock.move', 'create', [{
                'name':             row.get('name', 'Producto'),
                'picking_id':       picking_id,
                'product_id':       pp_id,
                'product_uom_qty':  qty,
                'product_uom':      uom_id,
                'location_id':      loc_src_id,
                'location_dest_id': loc_dest_id,
            }], {})

        # 5. Confirmar, asignar y validar
        odoo_call('stock.picking', 'action_confirm', [[picking_id]], {})
        odoo_call('stock.picking', 'action_assign',  [[picking_id]], {})
        try:
            odoo_call('stock.picking', 'button_validate', [[picking_id]], {})
            state = 'done'
        except Exception as e:
            state = 'confirmed'
            print(f'  [TRNF] Validación parcial: {e}')

        pick_data = odoo_call('stock.picking', 'read', [[picking_id]], {'fields': ['name', 'state']})
        pick_name = pick_data[0]['name'] if pick_data else str(picking_id)

        return {'ok': True, 'picking_id': picking_id, 'picking_name': pick_name,
                'state': state, 'count': len(rows)}

    def log_message(self, fmt, *args):
        print(f'  [{self.address_string()}] {fmt % args}')


# ── Main ──────────────────────────────────────────────────────
if __name__ == '__main__':
    # Siempre servir archivos desde la carpeta del proyecto
    import os
    _base_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(_base_dir)

    port = int(os.environ.get("PORT", 8080))
    print(f'\n🔥 Tecnigass – Recepción/Catálogo corriendo en puerto {port}')
    print(f'   Dashboard:   http://localhost:{port}/dashboard.html')
    print(f'   Catálogo:    http://localhost:{port}/index.html')
    print(f'   Recepción:   http://localhost:{port}/recepcion.html')
    print(f'   Traslado:    http://localhost:{port}/traslado.html')
    print('   Presiona Ctrl+C para detener\n')
    HTTPServer(('', port), Handler).serve_forever()
